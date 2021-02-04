from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from selenium import webdriver
import openpyxl
import csv
import smtplib, ssl
from mailer import Mailer
import os
import smtplib
import pandas as pd
import imghdr
from email.message import EmailMessage
from unidecode import unidecode
from datetime import datetime
import xlrd

driverPath = "drivers/geckodriver.exe"  # Insert your drivers path C://Windows
mime_types = "application/pdf,application/vnd.adobe.xfdf,application/vnd.fdf,application/vnd.adobe.xdp+xml"

fp = webdriver.FirefoxProfile()
fp.set_preference("browser.download.folderList", 2)
fp.set_preference("browser.download.manager.showWhenStarting", False)
fp.set_preference("plugin.disable_full_page_plugin_for_types", "application/pdf")
fp.set_preference("browser.helperApps.neverAsk.saveToDisk", mime_types)
fp.set_preference("plugin.disable_full_page_plugin_for_types", mime_types)
fp.set_preference("pdfjs.disabled", True)
fp.set_preference("javascript.enabled", False)

options = webdriver.FirefoxOptions()
# options.headless = True
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-notifications")
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')


def getInfoFromCsv():
    data = []
    wb = xlrd.open_workbook("Control.xls")
    sheet = wb.sheet_by_index(0)
    for i in range(sheet.nrows):
        data.append(sheet.cell(i, 1).value)
    # print(data)
    return data


def generateList(dataList):
    with open(outputPath + "output.csv", "r") as csvFile:
        reader = csv.reader(csvFile)
        for row in reader:
            dataList.append(row)
    return dataList


def getElements(row):
    x = fillingDate = tradeDate = companyName = insider = title = tradeType = price = qty = owned = detaOwn = value = d1 = w1 = m1 = m = ""
    i = 0
    for cell in row.find_elements_by_tag_name("td"):
        s = str(cell.text).strip()
        if s == "":
            s = " "
        if i == 0:
            x = s
            i += 1
            continue
        if i == 1:
            fillingDate = s
            i += 1
            continue
        if i == 2:
            tradeDate = s
            i += 1
            continue
        if i == 3:
            ticker = s
            i += 1
            continue
        if i == 4:
            companyName = s
            i += 1
            continue
        if i == 5:
            insider = s
            i += 1
            continue
        if i == 6:
            title = s
            i += 1
            continue
        if i == 7:
            tradeType = s
            i += 1
            continue
        if i == 8:
            price = s
            i += 1
            continue
        if i == 9:
            qty = s
            i += 1
            continue
        if i == 10:
            owned = s
            i += 1
            continue
        if i == 11:
            detaOwn = s
            i += 1
            continue
        if i == 12:
            value = s
            i += 1
            continue
        if i == 13:
            d1 = s
            i += 1
            continue
        if i == 14:
            w1 = s
            i += 1
            continue
        if i == 15:
            m1 = s
            i += 1
            continue
        if i == 16:
            m6 = s
            i = 0
            continue
    # print([fillingDate, tradeDate, ticker, insider, title, tradeType, price, qty, owned, detaOwn, value, x])
    return x, fillingDate, tradeDate, ticker, companyName, insider, title, tradeType, price, qty, owned, detaOwn, value, d1, w1, m1, m6


def checkInCsv(subjectParam, dataList, x, fillingDate, tradeDate, ticker, companyName, insider, title, tradeType, price,
               qty, owned,
               detaOwn, value, d1, w1, m1, m6):
    for ele in dataList:
        if ele[0] == fillingDate and ele[1] == tradeDate and ele[2] == ticker and ele[3] == insider and ele[4] == title:
            return True
    print([fillingDate, tradeDate, ticker, insider, title, tradeType, price, qty, owned, detaOwn, value, x])
    dataList.insert(0,
                    [fillingDate, tradeDate, ticker, insider, title, tradeType, price, qty, owned, detaOwn, value, x])

    success, excelData, excelCols = valid(fillingDate, tradeDate, ticker, insider, title, tradeType, price, qty, owned,
                                          detaOwn,
                                          value, x)
    if success and os.path.isfile(outputPath + "output.csv"):
        sendMail(excelData, dataList[0], subjectParam, excelCols)
    return False


def valid(fillingDate, tradeDate, ticker, insider, title, tradeType, price, qty, owned, detaOwn, value, x):
    excelData = []
    excelCols = []
    tickerCheck = False
    volumeTCheck = False
    priceCheck = False

    wb = xlrd.open_workbook(excelPath + "/" + excelName)
    sheet = wb.sheet_by_index(0)
    for i in range(1, sheet.nrows):
        if ticker == str(sheet.cell(i, 0).value):
            # print(ticker, str(sheet.cell(i, 0).value))
            tickerCheck = True
            if float(sheet.cell(i, volumeColumn - 1).value) >= minVol:
                volumeTCheck = True
            if float(price.replace("$", "").replace(",", "")) >= (1-discount/100) * float(sheet.cell(i, priceColumn-1).value):
                priceCheck = True
            for j in range(0, sheet.ncols):
                if str(sheet.cell(0, j).value) == "":
                    excelCols.append("column" + str(j))
                else:
                    excelCols.append(str(sheet.cell(0, j).value))
                if str(sheet.cell(i, j).value) == "":
                    excelData.append("-")
                else:
                    excelData.append(str(sheet.cell(i, j).value))
        if tickerCheck:
            break
    if tickerCheck and volumeTCheck and priceCheck:
        return True, excelData, excelCols
    return False, excelData, excelCols


def sendMail(excelData, arr, subjectParam, excelCols):
    if subjectParam == 0:
        SUBJECT = "Officer, "
    else:
        SUBJECT = "Other, "
    SUBJECT += arr[2] + ", " + arr[10] + ", " + arr[11] + ", " + excelData[6] + ", " + excelData[16] + ", " + excelData[17] + ", " + excelData[29]
    SUBJECT= From+", "+SUBJECT
    if os.path.exists("Table1.html"):
        os.remove("Table1.html")
    if os.path.exists("Table2.html"):
        os.remove("Table2.html")

    df1 = pd.DataFrame(data=[arr],
                       columns=["Filling Date", "Trade Date", "Ticker", "Insider Name", "Title", "Trade Type", "Price",
                                "Qty", "Owned", "Own Inc", "Value", "X"])
    df1.to_html("Table.html", index=False)
    htmltable1 = df1.to_html(index=False)

    df2 = pd.DataFrame(data=[excelData], columns=excelCols)
    df2.to_html("Table2.html", index=False)
    htmltable2 = df2.to_html(index=False)

    HTML = """
        <!DOCTYPE html>
        <html>
          <head>
            <meta charset="utf-8" />
            <style type="text/css">
              table {
                background: white;
                border-radius:3px;
                border-collapse: collapse;
                height: auto;
                max-width: 900px;
                padding:5px;
                width: 100%;
                animation: float 5s infinite;
              }
              th {
                color:#D5DDE5;;
                background:#1b1e24;
                border-bottom: 4px solid #9ea7af;
                font-size:14px;
                font-weight: 300;
                padding:10px;
                text-align:center;
                vertical-align:middle;
              }
              tr {
                border-top: 1px solid #C1C3D1;
                border-bottom: 1px solid #C1C3D1;
                border-left: 1px solid #C1C3D1;
                color:#666B85;
                font-size:16px;
                font-weight:normal;
              }
              tr:hover td {
                background:#4E5066;
                color:#FFFFFF;
                border-top: 1px solid #22262e;
              }
              td {
                background:#FFFFFF;
                padding:10px;
                text-align:left;
                vertical-align:middle;
                font-weight:300;
                font-size:13px;
                border-right: 1px solid #C1C3D1;
              }
            </style>
          </head>
          <body>""" + htmltable1 + """<br>""" + htmltable2 + """</body>
        </html>
        """

    SERVER = str(smtpServeName) + ":" + str(int(smtpServerPort))
    FROM = senderEmailId
    TO = receiverEmail

    message = MIMEMultipart('alternative')
    part1 = MIMEText(HTML, "html")
    message.attach(part1)
    message['Subject'] = SUBJECT
    message['From'] = FROM
    message['To'] = TO
    server = smtplib.SMTP(SERVER)
    if smtpServeName=="smtp.gmail.com":
        server.starttls()
        server.login(senderEmailId, str(senderEmailPassword))  # use str(int(pass)) for is password in numeric only

    else:
        server.login(senderEmailId, str(int(senderEmailPassword)))
    server.sendmail(FROM, TO, str(message))
    server.quit()
    print("email sent")


value1, value2, minVol, discount, smtpServeName, \
smtpServerPort, senderEmailId, senderEmailPassword, \
secure, encryption, receiverEmail, From, excelPath, \
excelName, outputPath, priceColumn, volumeColumn, sleep = getInfoFromCsv()
outputPath += '\\'
excelPath += '\\'

value1=int(value1)
value2=int(value2)
discount=float(discount)
smtpServerPort=int(smtpServerPort)
priceColumn=int(priceColumn)
volumeColumn=int(volumeColumn)
if str(senderEmailPassword).isnumeric():
    senderEmailPassword=int(senderEmailPassword)
# senderEmailPassword=int(senderEmailPassword)
